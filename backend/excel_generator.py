"""
Excel Generation Module
Creates a single Excel file with 8 sheets from extracted PDF data
File: backend/excel_generator.py

CRITICAL FIXES:
- "Contact No." with period for Buyer and Service Provider
- "Type of car" with lowercase 'c'
- NA preservation (do NOT filter out)
"""

from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """
    Production-grade Excel generator for GeM contract extraction.

    Guarantees:
    - Strict schema-to-column mapping
    - Deterministic column order
    - English-only output
    - EMPTY values remain blank
    - NA values are PRESERVED (not converted to blank)
    - No fuzzy or heuristic matching
    """

    MAX_SHEET_NAME_LENGTH = 31
    INVALID_SHEET_CHARS = ["\\", "/", "*", "?", ":", "[", "]"]

    def __init__(self) -> None:
        self.header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    # =====================================================
    # Public API
    # =====================================================

    def generate_excel(
        self,
        extracted_data: List[Dict[str, Any]],
        output_path: Path,
    ) -> None:
        if not isinstance(extracted_data, list):
            raise ValueError("extracted_data must be a list of dictionaries")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        try:
            self._sheet_contract_data(wb, extracted_data)
            self._sheet_organisation_details(wb, extracted_data)
            self._sheet_buyer_details(wb, extracted_data)
            self._sheet_financial_approval(wb, extracted_data)
            self._sheet_paying_authority(wb, extracted_data)
            self._sheet_consignee_details(wb, extracted_data)
            self._sheet_service_provider(wb, extracted_data)
            self._sheet_service_details(wb, extracted_data)

            wb.save(output_path)

        finally:
            wb.close()

    # =====================================================
    # Sheet Builders (STRICT, SCHEMA-LOCKED)
    # =====================================================

    def _sheet_contract_data(self, wb, data):
        headers = [
            "Contract No",
            "Contract Generated Date",
            "Bid / RA / PBP No",
            "Duration",
            "Amount of Contract (Including All Duties and Taxes INR)",
        ]

        ws = wb.create_sheet(self._sanitize_sheet_name("Contract Data"))
        self._write_headers(ws, headers)

        for row, item in enumerate(data, start=2):
            section = self._safe_section(item, "contract_data")
            self._write_row(ws, row, [self._safe_get(section, h) for h in headers])

        self._auto_adjust_columns(ws)

    def _sheet_organisation_details(self, wb, data):
        headers = ["Type", "Ministry", "Department", "Organisation Name", "Office Zone"]
        ws = wb.create_sheet(self._sanitize_sheet_name("Organisation Details"))
        self._write_headers(ws, headers)

        for row, item in enumerate(data, start=2):
            section = self._safe_section(item, "organisation_details")
            self._write_row(ws, row, [self._safe_get(section, h) for h in headers])

        self._auto_adjust_columns(ws)

    def _sheet_buyer_details(self, wb, data):
        """
        CRITICAL FIX: "Contact No." with period to match schema alias
        """
        headers = ["Designation", "Contact No.", "Email ID", "GSTIN", "Address"]
        #                         ^^^^^^^^^^^^
        #                         ADDED PERIOD!
        
        ws = wb.create_sheet(self._sanitize_sheet_name("Buyer Details"))
        self._write_headers(ws, headers)

        for row, item in enumerate(data, start=2):
            section = self._safe_section(item, "buyer_details")
            self._write_row(ws, row, [self._safe_get(section, h) for h in headers])

        self._auto_adjust_columns(ws)

    def _sheet_financial_approval(self, wb, data):
        headers = [
            "IFD Concurrence",
            "Designation of Administrative Approval",
            "Designation of Financial Approval",
        ]
        ws = wb.create_sheet(self._sanitize_sheet_name("Financial Approval Details"))
        self._write_headers(ws, headers)

        for row, item in enumerate(data, start=2):
            section = self._safe_section(item, "financial_approval_details")
            self._write_row(ws, row, [self._safe_get(section, h) for h in headers])

        self._auto_adjust_columns(ws)

    def _sheet_paying_authority(self, wb, data):
        headers = ["Role", "Payment Mode", "Designation", "Email ID", "GSTIN", "Address"]
        ws = wb.create_sheet(self._sanitize_sheet_name("Paying Authority Details"))
        self._write_headers(ws, headers)

        for row, item in enumerate(data, start=2):
            section = self._safe_section(item, "paying_authority_details")
            self._write_row(ws, row, [self._safe_get(section, h) for h in headers])

        self._auto_adjust_columns(ws)

    def _sheet_consignee_details(self, wb, data):
        headers = ["Contact", "Email ID", "GSTIN", "Address", "Service Description"]
        ws = wb.create_sheet(self._sanitize_sheet_name("Consignee Details"))
        self._write_headers(ws, headers)

        for row, item in enumerate(data, start=2):
            section = self._safe_section(item, "consignee_details")
            self._write_row(ws, row, [self._safe_get(section, h) for h in headers])

        self._auto_adjust_columns(ws)

    def _sheet_service_provider(self, wb, data):
        """
        CRITICAL FIX: "Contact No." with period to match schema alias
        """
        headers = [
            "GeM Seller ID",
            "Company Name",
            "Contact No.",  # ✅ ADDED PERIOD!
            "Email ID",
            "Address",
            "MSME Registration Number",
            "GSTIN",
            "MSME Status as verified by buyer",
            "MSE Social Category",
            "MSE Gender",
        ]
        ws = wb.create_sheet(self._sanitize_sheet_name("Service Provider Details"))
        self._write_headers(ws, headers)

        for row, item in enumerate(data, start=2):
            section = self._safe_section(item, "service_provider_details")
            self._write_row(ws, row, [self._safe_get(section, h) for h in headers])

        self._auto_adjust_columns(ws)

    def _sheet_service_details(self, wb, data):
        """
        CRITICAL FIX: "Type of car" with lowercase 'c' to match schema alias
        """
        headers = [
            "Service Start Date (latest by)",
            "Service End Date",
            "Category Name",
            "Billing Cycle",
            "District",
            "Zipcode",
            "Vehicle Type",
            "Type of car (Please select at least 3 options)",  # ✅ LOWERCASE 'c'!
        ]
        ws = wb.create_sheet(self._sanitize_sheet_name("Service Details"))
        self._write_headers(ws, headers)

        for row, item in enumerate(data, start=2):
            section = self._safe_section(item, "service_details")
            self._write_row(ws, row, [self._safe_get(section, h) for h in headers])

        self._auto_adjust_columns(ws)

    # =====================================================
    # Helpers
    # =====================================================

    @staticmethod
    def _safe_section(item: Dict[str, Any], key: str) -> Dict[str, Any]:
        section = item.get(key, {})
        return section if isinstance(section, dict) else {}

    @staticmethod
    def _safe_get(section: Dict[str, Any], key: str) -> str:
        """
        CRITICAL FIX: Preserve "NA" values - do NOT convert to blank!
        
        Original logic was:
            if value in (None, "", "NA", "-"):
                return ""
        
        This was WRONG because it converts "NA" to blank.
        
        New logic:
        - None → ""
        - "" → ""
        - "-" → "" (hyphen is a placeholder)
        - "NA" → "NA" (PRESERVED!)
        - Any other value → str(value)
        """
        value = section.get(key)
        
        # Only filter out None, empty string, and hyphen
        if value in (None, "", "-"):
            return ""
        
        # Preserve everything else including "NA"
        return str(value)

    def _sanitize_sheet_name(self, title: str) -> str:
        for ch in self.INVALID_SHEET_CHARS:
            title = title.replace(ch, "-")
        return (title.strip() or "Sheet")[: self.MAX_SHEET_NAME_LENGTH]

    def _write_headers(self, ws, headers):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    def _write_row(self, ws, row, values):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    def _auto_adjust_columns(self, ws):
        for column in ws.columns:
            col_letter = get_column_letter(column[0].column)
            max_len = max(
                (len(str(cell.value)) for cell in column if cell.value),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)